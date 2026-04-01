# -*- coding: utf-8 -*-
"""
Version NUBE del scraping - corre en GitHub Actions.
Solo usa la API (sin Playwright) para ser rapido y ligero.
Los datos de cronograma se obtienen del estado de la API.
"""

import urllib.request, urllib.parse, json, re, os, smtplib, time
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

API_BASE   = "https://dev.activosporcolombia.com/net/api"
SITE_BASE  = "https://activosporcolombia.com"
CITY_ID_MEDELLIN  = 5001
DIAS_ROJO  = 2

# Credenciales desde variables de entorno (GitHub Secrets)
EMAIL_REMITENTE    = os.environ.get("EMAIL_REMITENTE", "")
EMAIL_CONTRASENA   = os.environ.get("GMAIL_APP_PASSWORD", "")
EMAIL_DESTINATARIO = os.environ.get("EMAIL_DESTINATARIO", "")

TIPOS = {
    1:"Apartaestudio",2:"Apartamento",3:"Bodega",4:"Casa",5:"Casa Lote",
    6:"Casa Recreo",7:"Centro Comercial",11:"Consultorio",13:"Hotel/Motel",
    14:"Edificio",16:"Finca",17:"Garaje",21:"Local Comercial",22:"Lote",
    24:"Lote con Construccion",29:"Oficina",30:"Parqueadero",
}

BARRIOS = {}  # se carga desde la API de filtros


def llamar_api(endpoint, params=None, reintentos=3):
    url = f"{API_BASE}{endpoint}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    for intento in range(reintentos):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.loads(r.read().decode("utf-8"))
        except Exception:
            if intento < reintentos - 1:
                time.sleep(3)
            else:
                raise


def cargar_barrios():
    """Carga mapa de neighborhood_id -> nombre desde la API de filtros."""
    global BARRIOS
    try:
        resp = llamar_api("/public/v1/inmuebles/filtros")
        for dept in resp["data"]["ubicacion"]["departamentos"]:
            for city in dept.get("cities", []):
                for barrio in city.get("neighborhoods", []):
                    BARRIOS[barrio["id"]] = barrio["nombre"]
    except Exception:
        pass


def obtener_propiedades(filtro):
    todos, pagina = [], 1
    while True:
        params = {"query":"","page":pagina,"limit":50,"sort_by":"date_desc"}
        params.update(filtro)
        resp = llamar_api("/v1/properties/search", params)
        data = resp["data"]
        props = data.get("properties", [])
        if not props: break
        todos.extend(props)
        if pagina >= data.get("total_pages", 1): break
        pagina += 1
    return todos


def construir_url(p):
    itype = (p.get("item_type") or "").upper()
    pid = p["id"]
    ref = (p.get("reference") or "").lower()
    slug = re.sub(r"[^a-z0-9\-]", "", ref.replace(" - ","-").replace(" ","-"))
    if itype == "UNIDAD_INMOBILIARIA":
        return f"{SITE_BASE}/es/unidad-inmobiliaria/{pid}/{slug}"
    return f"{SITE_BASE}/es/inmueble/{pid}/{slug}"


def formatear_precio(v):
    if not v: return "X"
    try: return "$ {:,.0f}".format(float(v)).replace(",",".")
    except: return str(v)


def procesar(props):
    resultado = []
    for p in props:
        estado = p.get("state") or {}
        estado_cod = estado.get("code","") if isinstance(estado, dict) else ""
        estado_nom = estado.get("name","") if isinstance(estado, dict) else str(estado)

        barrio = BARRIOS.get(p.get("neighborhood_id"), "")
        nombre = barrio or (p.get("reference") or "").split(" - ")[0].strip()

        # Determinar cronograma desde estado de la API
        if "PROXIMO" in estado_cod.upper() or "SUBASTA" in estado_cod.upper():
            estado_crono = "Con cronograma"
            etapa = estado_nom
        elif "MANIFEST" in estado_cod.upper() or "ABIERTA" in estado_cod.upper():
            estado_crono = "Manifestacion Abierta"
            etapa = ""
        else:
            estado_crono = estado_nom or "Sin info"
            etapa = ""

        resultado.append({
            "_id": str(p["id"]),
            "nombre": nombre,
            "direccion": f"{barrio}, {p.get('location','')}" if barrio else p.get("location",""),
            "tipo": TIPOS.get(p.get("property_type_id"), ""),
            "area_m2": p.get("built_area") or p.get("lot_area") or "",
            "valor": formatear_precio(p.get("base_sale_price") or p.get("commercial_appraisal")),
            "estado_crono": estado_crono,
            "etapa_actual": etapa,
            "plazo": "X",
            "estado_api": estado_cod,
            "link": construir_url(p),
        })
    return resultado


# ── Deteccion de cambios ──

CAMPOS = ["nombre","direccion","tipo","valor","estado_crono","etapa_actual","plazo"]
DATOS_FILE = os.path.join(os.path.dirname(__file__), "datos_anteriores.json")
CAMBIOS_FILE = os.path.join(os.path.dirname(__file__), "registro_cambios.json")


def cargar_json(path):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def detectar_cambios(inmuebles, tab):
    anteriores = cargar_json(DATOS_FILE)
    registro = cargar_json(CAMBIOS_FILE)
    ahora = datetime.now().isoformat()
    limite = (datetime.now() - timedelta(days=DIAS_ROJO)).isoformat()
    resumen = []

    for item in inmuebles:
        pid = item["_id"]
        base = f"{tab}:{pid}"
        prev = anteriores.get(base, {})
        for campo in CAMPOS:
            nuevo = str(item.get(campo,""))
            viejo = str(prev.get(campo,""))
            if prev and nuevo != viejo:
                registro[f"{base}:{campo}"] = ahora
                resumen.append(f"[{tab}] {item.get('nombre','?')}: {campo} '{viejo}' -> '{nuevo}'")
        anteriores[base] = {c: str(item.get(c,"")) for c in CAMPOS}

    for k in list(registro.keys()):
        if registro[k] < limite:
            del registro[k]

    guardar_json(DATOS_FILE, anteriores)
    guardar_json(CAMBIOS_FILE, registro)
    return {k:True for k,v in registro.items() if v >= limite}, resumen


# ── Excel ──

def guardar_excel(med, ant, cambios_med, cambios_ant):
    import openpyxl, openpyxl.styles, openpyxl.utils

    COLOR_HEADER="1F3864"; COLOR_CRONO="D6E4F0"; COLOR_MANIF="F2F2F2"
    COLOR_SUB="FFF2CC"; COLOR_CAMBIO="FF4444"; FONT_W="FFFFFF"

    def color_fila(ec, ea):
        cod = (ea or "").upper()
        if ec == "Manifestacion Abierta": return COLOR_MANIF
        if "SUBASTA" in cod or "PROXIMO" in cod: return COLOR_SUB
        return COLOR_CRONO

    def escribir(wb, titulo_hoja, titulo, items, cambios, tab):
        ws = wb.create_sheet(titulo_hoja)
        COLS=["NOMBRE","DIRECCION","TIPO","AREA m2","VALOR","ESTADO CRONOGRAMA","ETAPA ACTUAL","PLAZO","CLIENTE","LINK"]
        CAMPOS_E=["nombre","direccion","tipo","area_m2","valor","estado_crono","etapa_actual","plazo","_c","link"]
        ANCHOS=[30,38,18,10,20,22,35,10,14,55]
        ALIN=["l","l","c","c","c","c","l","c","c","l"]
        borde=openpyxl.styles.Border(*(openpyxl.styles.Side("thin",color="CCCCCC") for _ in range(4)))

        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=len(COLS))
        t=ws.cell(row=1,column=1,value=titulo)
        t.font=openpyxl.styles.Font(name="Calibri",bold=True,size=14,color=FONT_W)
        t.fill=openpyxl.styles.PatternFill("solid",fgColor=COLOR_HEADER)
        t.alignment=openpyxl.styles.Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height=28

        for c,n in enumerate(COLS,1):
            cl=ws.cell(row=2,column=c,value=n)
            cl.fill=openpyxl.styles.PatternFill("solid",fgColor="2E5D9E")
            cl.font=openpyxl.styles.Font(name="Calibri",bold=True,color=FONT_W,size=10)
            cl.alignment=openpyxl.styles.Alignment(horizontal="center",vertical="center",wrap_text=True)
            cl.border=borde
        ws.row_dimensions[2].height=30

        rojo_fill=openpyxl.styles.PatternFill("solid",fgColor=COLOR_CAMBIO)
        rojo_font=openpyxl.styles.Font(name="Calibri",size=10,color="FFFFFF",bold=True)

        for fila,item in enumerate(items,3):
            pid=item["_id"]
            cf=color_fila(item.get("estado_crono",""),item.get("estado_api",""))
            base_fill=openpyxl.styles.PatternFill("solid",fgColor=cf)
            base_font=openpyxl.styles.Font(name="Calibri",size=10)
            for c,(campo,al) in enumerate(zip(CAMPOS_E,ALIN),1):
                val="Grupo NBC" if campo=="_c" else item.get(campo,"")
                cl=ws.cell(row=fila,column=c,value=val)
                ck=f"{tab}:{pid}:{campo}"
                if ck in cambios: cl.fill=rojo_fill; cl.font=rojo_font
                else: cl.fill=base_fill; cl.font=base_font
                cl.alignment=openpyxl.styles.Alignment(horizontal="left" if al=="l" else "center",vertical="center",wrap_text=(al=="l"))
                cl.border=borde
            ws.row_dimensions[fila].height=22

        for c,w in enumerate(ANCHOS,1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width=w
        ws.auto_filter.ref=f"A2:{openpyxl.utils.get_column_letter(len(COLS))}{len(items)+2}"
        ws.freeze_panes="A3"

    wb=openpyxl.Workbook()
    wb.remove(wb.active)
    escribir(wb,"Medellin","LISTADO DE VENTA MASIVA - MEDELLIN",med,cambios_med,"med")
    escribir(wb,"Antioquia","LISTADO DE VENTA MASIVA - ANTIOQUIA (sin Medellin)",ant,cambios_ant,"ant")

    wl=wb.create_sheet("Leyenda")
    wl["A1"]="Leyenda de colores"; wl["A1"].font=openpyxl.styles.Font(bold=True)
    for i,(d,co) in enumerate([("Proximo Subasta",COLOR_SUB),("En proceso",COLOR_CRONO),("Manifestacion Abierta",COLOR_MANIF),("DATO CAMBIO (rojo 2 dias)",COLOR_CAMBIO)],2):
        c=wl.cell(row=i,column=1,value=d); c.fill=openpyxl.styles.PatternFill("solid",fgColor=co)
        if co==COLOR_CAMBIO: c.font=openpyxl.styles.Font(color="FFFFFF",bold=True)
    wl.column_dimensions["A"].width=40

    wi=wb.create_sheet("Info")
    wi["A1"]="Actualizacion"; wi["B1"]=datetime.now().strftime("%Y-%m-%d %H:%M")
    wi["A2"]="Medellin"; wi["B2"]=len(med)
    wi["A3"]="Antioquia"; wi["B3"]=len(ant)

    out = os.path.join(os.path.dirname(__file__), "inmuebles_medellin.xlsx")
    wb.save(out)
    return out


# ── Email ──

def enviar_email(resumen):
    if not EMAIL_REMITENTE or not EMAIL_CONTRASENA or not resumen:
        return
    try:
        cuerpo = f"Se detectaron {len(resumen)} cambios:\n\n"
        cuerpo += "\n".join(f"  - {c}" for c in resumen[:30])
        if len(resumen)>30: cuerpo += f"\n  ... y {len(resumen)-30} mas"
        cuerpo += f"\n\nFecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        msg = MIMEMultipart()
        msg["From"]=EMAIL_REMITENTE; msg["To"]=EMAIL_DESTINATARIO
        msg["Subject"]=f"Activos Colombia - {len(resumen)} cambios detectados"
        msg.attach(MIMEText(cuerpo,"plain"))

        with smtplib.SMTP_SSL("smtp.gmail.com",465) as s:
            s.login(EMAIL_REMITENTE, EMAIL_CONTRASENA)
            s.send_message(msg)
        print(f"Email enviado a {EMAIL_DESTINATARIO}")
    except Exception as e:
        print(f"Error email: {e}")


# ── Main ──

if __name__ == "__main__":
    print("Scraping NUBE - activosporcolombia.com")
    print("="*45)

    cargar_barrios()

    print("Descargando Medellin...")
    props_med = obtener_propiedades({"city_ids": CITY_ID_MEDELLIN})
    print(f"  {len(props_med)} propiedades")

    print("Descargando Antioquia (sin Medellin)...")
    todas = obtener_propiedades({})
    props_ant = [p for p in todas if p.get("city_id") and 5000<=p["city_id"]<=5999 and p["city_id"]!=CITY_ID_MEDELLIN]
    print(f"  {len(props_ant)} propiedades")

    med = procesar(props_med)
    ant = procesar(props_ant)

    cm, rm = detectar_cambios(med, "med")
    ca, ra = detectar_cambios(ant, "ant")
    cambios = rm + ra

    if cambios:
        print(f"\n*** {len(cambios)} CAMBIOS ***")
        for c in cambios[:10]: print(f"  - {c}")
    else:
        print("Sin cambios")

    archivo = guardar_excel(med, ant, cm, ca)
    print(f"Excel: {archivo}")

    if cambios:
        enviar_email(cambios)

    print(f"\nListo: {len(med)} Medellin + {len(ant)} Antioquia")
