# -*- coding: utf-8 -*-
import json, re, urllib.request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

API = "https://dev.activosporcolombia.com/net/api/v1/properties/search"
SITE_BASE = "https://activosporcolombia.com"
TIPOS = {
    1:"Apartaestudio",2:"Apartamento",3:"Bodega",4:"Casa",5:"Casa Lote",
    6:"Casa Recreo",7:"Centro Comercial",11:"Consultorio",13:"Hotel/Motel",
    14:"Edificio",16:"Finca",17:"Garaje",21:"Local Comercial",22:"Lote",
    24:"Lote con Construccion",29:"Oficina",30:"Parqueadero",
    12:"Deposito",18:"Habitacion Hotel",19:"Hacienda",35:"Avaluo",
}

def construir_url(p):
    pid = p["id"]
    ref = (p.get("reference") or "").lower()
    slug = re.sub(r"[^a-z0-9\-]", "", ref.replace(" - ","-").replace(" ","-"))
    return f"{SITE_BASE}/es/inmueble/{pid}/{slug}"

def fmt_precio(v):
    if not v: return ""
    try: return "$ {:,.0f}".format(float(v)).replace(",",".")
    except: return str(v)

def prop_to_row(p, seccion=""):
    estado = ""
    if isinstance(p.get("state"), dict):
        estado = (p["state"] or {}).get("name","")
    return [
        p.get("reference",""),
        "",
        TIPOS.get(p.get("property_type_id"),""),
        p.get("matricula_number",""),
        estado,
        "", "",
        "Grupo NBC",
        construir_url(p),
        str(p.get("built_area") or p.get("lot_area") or ""),
        fmt_precio(p.get("base_sale_price") or p.get("commercial_appraisal")),
        p.get("matricula_number",""),
        seccion,
    ]

HEADERS = ["NOMBRE","DIRECCION","TIPO","FOLIO MATRICULA","ESTADO","ETAPA ACTUAL",
           "PLAZO","CLIENTE","LINK","AREA m2","VALOR","FMI","SECCION"]

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
EDS_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GARAJE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_section_header(ws, row, text, fill=None):
    ws.cell(row=row, column=1, value=text).font = SECTION_FONT
    f = fill or SECTION_FILL
    for j in range(1, len(HEADERS)+1):
        ws.cell(row=row, column=j).fill = f

def write_row(ws, row, vals):
    for j, v in enumerate(vals, 1):
        ws.cell(row=row, column=j, value=v).border = BORDER

# Load data
with open('_temp_Bello.json','r',encoding='utf-8') as f:
    bello_data = json.load(f)
with open('_temp_La_Pintada.json','r',encoding='utf-8') as f:
    pintada_data = json.load(f)
with open('_temp_links.json','r',encoding='utf-8') as f:
    links_data = json.load(f)
with open('_temp_barrios.json','r',encoding='utf-8') as f:
    barrios_data = json.load(f)
with open('_temp_garajes_metro.json','r',encoding='utf-8') as f:
    garajes_metro = json.load(f)

# Caceres + El Bagre
caceres_bagre = []
for cid, nombre in [(5120, "Caceres"), (5250, "El Bagre")]:
    url = f"{API}?per_page=100&city_ids={cid}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as r:
        data = json.loads(r.read())
    for p in data["data"]["properties"]:
        caceres_bagre.append((p, nombre))

# ══════════════════════════════════════════════════════════════════
# ARCHIVO 1
# ══════════════════════════════════════════════════════════════════
wb1 = Workbook()
ws1 = wb1.active
ws1.title = "Inmuebles Especiales"

for j, h in enumerate(HEADERS, 1):
    cell = ws1.cell(row=1, column=j, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER

r = 2

# BELLO
write_section_header(ws1, r, "BELLO")
r += 1
for row in bello_data[2:]:
    if not any(row): continue
    vals = [
        row[0] if len(row)>0 else "",
        row[1] if len(row)>1 else "",
        row[2] if len(row)>2 else "",
        "",
        row[3] if len(row)>3 else "",
        row[4] if len(row)>4 else "",
        row[5] if len(row)>5 else "",
        row[6] if len(row)>6 else "",
        row[7] if len(row)>7 else "",
        row[8] if len(row)>8 else "",
        row[9] if len(row)>9 else "",
        row[10] if len(row)>10 else "",
        "Bello",
    ]
    write_row(ws1, r, vals)
    r += 1

# LA PINTADA
r += 1
write_section_header(ws1, r, "LA PINTADA")
r += 1
for row in pintada_data[2:]:
    if not any(row): continue
    vals = [
        row[0] if len(row)>0 else "",
        row[1] if len(row)>1 else "",
        row[2] if len(row)>2 else "",
        "",
        row[3] if len(row)>3 else "",
        row[4] if len(row)>4 else "",
        row[5] if len(row)>5 else "",
        row[6] if len(row)>6 else "",
        row[7] if len(row)>7 else "",
        row[8] if len(row)>8 else "",
        row[9] if len(row)>9 else "",
        row[10] if len(row)>10 else "",
        "La Pintada",
    ]
    write_row(ws1, r, vals)
    r += 1

# CACERES
r += 1
write_section_header(ws1, r, "CACERES")
r += 1
for p, ciudad in caceres_bagre:
    if p.get("city_id") == 5120:
        write_row(ws1, r, prop_to_row(p, "Caceres"))
        r += 1

# EL BAGRE
r += 1
write_section_header(ws1, r, "EL BAGRE")
r += 1
for p, ciudad in caceres_bagre:
    if p.get("city_id") == 5250:
        write_row(ws1, r, prop_to_row(p, "El Bagre"))
        r += 1

# EDS
r += 1
write_section_header(ws1, r, "ESTACIONES DE SERVICIO", EDS_FILL)
r += 1
eds_ids = ["14103", "13688", "11178"]
for pid_str in eds_ids:
    if pid_str in links_data:
        write_row(ws1, r, prop_to_row(links_data[pid_str], "EDS"))
        r += 1

# LOCAL CRA 70
r += 1
write_section_header(ws1, r, "LOCAL CRA 70 - MEDELLIN")
r += 1
if "467" in links_data:
    write_row(ws1, r, prop_to_row(links_data["467"], "Local Cra 70"))
    r += 1

# Format
widths = [35,35,18,18,20,30,10,15,55,10,18,18,15]
from openpyxl.utils import get_column_letter
for j, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(j)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{r-1}"

output1 = r"D:\Victoria\Trabajo\Papá\Trabajo\VS CODE\scraping-activos\Inmuebles_Especiales_v2.xlsx"
wb1.save(output1)
print(f"Archivo 1: {output1} ({r-2} filas)")

# ══════════════════════════════════════════════════════════════════
# ARCHIVO 2
# ══════════════════════════════════════════════════════════════════
wb2 = Workbook()
ws2 = wb2.active
ws2.title = "Laureles Belen Garajes"

for j, h in enumerate(HEADERS, 1):
    cell = ws2.cell(row=1, column=j, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER

r = 2

# Casas/Aptos
write_section_header(ws2, r, "CASAS Y APARTAMENTOS - LAURELES / BELEN / SIMON BOLIVAR")
r += 1
for item in barrios_data:
    vals = [
        item.get("nombre",""), item.get("direccion",""),
        item.get("tipo",""), item.get("matricula",""),
        item.get("estado",""), "", "",
        "Grupo NBC", item.get("link",""),
        item.get("area",""), item.get("precio",""),
        item.get("matricula",""), item.get("barrio",""),
    ]
    write_row(ws2, r, vals)
    r += 1

if not barrios_data:
    ws2.cell(row=r, column=1, value="(Solo 2 propiedades encontradas - ver nota abajo)").font = Font(italic=True, color="999999")
    r += 1

# Garajes
r += 1
write_section_header(ws2, r, "GARAJES - AREA METROPOLITANA", GARAJE_FILL)
r += 1

CITY_NAMES = {5001:"Medellin",5266:"Envigado",5631:"Sabaneta",5360:"Itagui",
              5088:"Bello",5212:"Copacabana",5129:"Caldas",5380:"La Estrella",
              5079:"Barbosa",5308:"Girardota"}

for p in garajes_metro:
    estado = ""
    if isinstance(p.get("state"), dict):
        estado = (p["state"] or {}).get("name","")
    ciudad = CITY_NAMES.get(p.get("city_id"), str(p.get("city_id","")))
    vals = [
        p.get("reference",""), "",
        TIPOS.get(p.get("property_type_id"),""),
        p.get("matricula_number",""),
        estado, "", "",
        "Grupo NBC", construir_url(p),
        str(p.get("built_area") or p.get("lot_area") or ""),
        fmt_precio(p.get("base_sale_price") or p.get("commercial_appraisal")),
        p.get("matricula_number",""), ciudad,
    ]
    write_row(ws2, r, vals)
    r += 1

for j, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{r-1}"

output2 = r"D:\Victoria\Trabajo\Papá\Trabajo\VS CODE\scraping-activos\Laureles_Belen_Garajes_v2.xlsx"
wb2.save(output2)
print(f"Archivo 2: {output2} ({r-2} filas)")
