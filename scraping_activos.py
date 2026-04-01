# -*- coding: utf-8 -*-
"""
Script de scraping - activosporcolombia.com
Extrae inmuebles en Medellin y Antioquia, detecta cambios y notifica.
Uso: py -X utf8 scraping_activos.py
"""

import urllib.request, urllib.parse, json, re, os, shutil, smtplib, time
import openpyxl, openpyxl.styles, openpyxl.utils
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from playwright.sync_api import sync_playwright

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIGURACION — editar aqui
# ═══════════════════════════════════════════════════════════════════════════════

CARPETA_SCRIPT     = r"D:\Victoria\Trabajo\Papá\Trabajo\VS CODE\scraping-activos"
ARCHIVO_EXCEL      = os.path.join(CARPETA_SCRIPT, "inmuebles_medellin.xlsx")
ARCHIVO_DRIVE      = r"G:\Mi unidad\inmuebles_medellin.xlsx"
ARCHIVO_DATOS_PREV = os.path.join(CARPETA_SCRIPT, "datos_anteriores.json")
ARCHIVO_CAMBIOS    = os.path.join(CARPETA_SCRIPT, "registro_cambios.json")

CITY_ID_MEDELLIN   = 5001
DEPT_ID_ANTIOQUIA  = 5
API_BASE           = "https://dev.activosporcolombia.com/net/api"
SITE_BASE          = "https://activosporcolombia.com"
DIAS_ROJO          = 2   # dias que un cambio se muestra en rojo

# Notificacion por correo (configurar para activar)
EMAIL_ACTIVADO     = True
EMAIL_REMITENTE    = "cuanticateamsas@gmail.com"
EMAIL_CONTRASENA   = os.environ.get("GMAIL_APP_PASSWORD", "etjd zvie ziib wcad")
EMAIL_DESTINATARIO = "cuanticateamsas@gmail.com"

TIPOS = {
    1:"Apartaestudio",2:"Apartamento",3:"Bodega",4:"Casa",5:"Casa Lote",
    6:"Casa Recreo",7:"Centro Comercial",11:"Consultorio",13:"Hotel/Motel",
    14:"Edificio",16:"Finca",17:"Garaje",21:"Local Comercial",22:"Lote",
    24:"Lote con Construccion",29:"Oficina",30:"Parqueadero",
}

# Colores
COLOR_HEADER     = "1F3864"
COLOR_CRONOGRAMA = "D6E4F0"
COLOR_MANIF      = "F2F2F2"
COLOR_SUBASTA    = "FFF2CC"
COLOR_CAMBIO     = "FF4444"    # Rojo para celdas con cambios
FONT_HEADER      = "FFFFFF"


# ═══════════════════════════════════════════════════════════════════════════════
# 1. API
# ═══════════════════════════════════════════════════════════════════════════════

def llamar_api(endpoint, params=None, reintentos=3):
    url = f"{API_BASE}{endpoint}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    for intento in range(reintentos):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.loads(r.read().decode("utf-8"))
        except Exception:
            if intento < reintentos - 1:
                time.sleep(3)
            else:
                raise


def obtener_propiedades(filtro_params):
    """Descarga todas las paginas de una busqueda."""
    todos, pagina = [], 1
    while True:
        params = {"query":"","page":pagina,"limit":50,"sort_by":"date_desc"}
        params.update(filtro_params)
        resp = llamar_api("/v1/properties/search", params)
        data = resp["data"]
        props = data.get("properties", [])
        if not props:
            break
        todos.extend(props)
        if pagina >= data.get("total_pages", 1):
            break
        pagina += 1
    return todos


def obtener_medellin():
    return obtener_propiedades({"city_ids": CITY_ID_MEDELLIN})


def obtener_antioquia_sin_medellin():
    # city_id de Antioquia en codigo DANE van de 5000 a 5999
    todas = obtener_propiedades({})  # todas las propiedades
    return [
        p for p in todas
        if p.get("city_id") and 5000 <= p["city_id"] <= 5999
        and p["city_id"] != CITY_ID_MEDELLIN
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# 2. PLAYWRIGHT — detalle de cada propiedad
# ═══════════════════════════════════════════════════════════════════════════════

def construir_url(p):
    itype = (p.get("item_type") or "").upper()
    pid   = p["id"]
    ref   = (p.get("reference") or "").lower()
    slug  = re.sub(r"[^a-z0-9\-]", "", ref.replace(" - ","-").replace(" ","-"))
    if itype == "UNIDAD_INMOBILIARIA":
        return f"{SITE_BASE}/es/unidad-inmobiliaria/{pid}/{slug}"
    return f"{SITE_BASE}/es/inmueble/{pid}/{slug}"


def resolver_ip():
    """Obtiene la IP actual de activosporcolombia.com via DNS de Google."""
    try:
        req = urllib.request.Request(
            "https://dns.google/resolve?name=activosporcolombia.com&type=A",
            headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as r:
            d = json.loads(r.read().decode("utf-8"))
            for a in d.get("Answer", []):
                if a.get("type") == 1:
                    return a["data"]
    except Exception:
        pass
    return "147.93.180.97"  # fallback


def scrape_detalles(propiedades, etiqueta=""):
    """Visita cada propiedad con Playwright y extrae nombre, direccion y cronograma."""
    ip = resolver_ip()
    resultados = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=[
            "--dns-prefetch-disable",
            f"--host-resolver-rules=MAP activosporcolombia.com {ip}",
            "--no-sandbox",
        ])
        page = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36"
        ).new_page()

        total = len(propiedades)
        for i, prop in enumerate(propiedades, 1):
            pid = str(prop["id"])
            url = construir_url(prop)
            ref = (prop.get("reference") or "")[:50]
            print(f"  {etiqueta}[{i}/{total}] {ref}")

            try:
                page.goto(url, wait_until="domcontentloaded", timeout=25000)
                page.wait_for_timeout(3000)
                lineas = [l.strip() for l in page.inner_text("body").split("\n") if l.strip()]
            except Exception:
                lineas = []

            direccion, barrio = "", ""
            for l in lineas:
                if l.startswith("Direcci"):
                    direccion = l.split(":", 1)[-1].strip()
                if l.startswith("Barrio:"):
                    barrio = l.split(":", 1)[-1].strip()

            nombre = extraer_nombre_edificio(lineas, barrio)
            estado_crono, etapa_actual, plazo = extraer_cronograma(lineas)

            resultados[pid] = {
                "nombre": nombre, "direccion": direccion, "barrio": barrio,
                "estado_crono": estado_crono, "etapa_actual": etapa_actual, "plazo": plazo,
            }

        browser.close()
    return resultados


def extraer_nombre_edificio(lineas, barrio_fallback):
    texto = " ".join(lineas)
    patrones = [
        r"(?:Edificio|Ed\.)\s+([A-Z\u00C0-\u00DC][A-Za-z\u00C0-\u00FC\u00F1\u00D1\s]+?)(?:\s*,|\s*\.|ubicad|situad)",
        r"(?:Conjunto Residencial|Conj\.)\s+([A-Z\u00C0-\u00DC][A-Za-z\u00C0-\u00FC\u00F1\u00D1\s]+?)(?:\s*,|\s*\.|ubicad|situad)",
        r"(?:Centro Comercial|CC)\s+([A-Z\u00C0-\u00DC][A-Za-z\u00C0-\u00FC\u00F1\u00D1\s]+?)(?:\s*,|\s*\.|ubicad|situad)",
        r"(?:Parque Empresarial|Torre|Local)\s+([A-Z\u00C0-\u00DC][A-Za-z\u00C0-\u00FC\u00F1\u00D1\s]+?)(?:\s*,|\s*\.|ubicad|situad)",
    ]
    for pat in patrones:
        m = re.search(pat, texto)
        if m:
            nombre = m.group(1).strip()
            if 3 < len(nombre) < 50:
                return nombre
    return barrio_fallback or "Sin nombre"


def extraer_cronograma(lineas):
    ETAPAS = [
        "PUBLICACI", "REGISTRO", "DILIGENCIA", "FINANCIERO",
        "CUPONES", "SERIEDAD", "VALIDACI", "SUBASTA",
    ]

    for l in lineas:
        if "manifestaci" in l.lower() and "abierta" in l.lower():
            return "Manifestacion Abierta", "", "X"

    if not any("ronograma" in l for l in lineas):
        return "Manifestacion Abierta", "", "X"

    etapa_activa = ""
    for i, l in enumerate(lineas):
        lu = l.upper()
        if any(e in lu for e in ETAPAS):
            contexto = " ".join(lineas[max(0,i-2):i+6]).upper()
            if "ACTIVO" in contexto:
                etapa_activa = l.strip()
                break

    plazo = "X"
    for l in lineas:
        m = re.search(r"Fin:\s*\w+,\s*(\d+)\s+de\s+(\w+)\s+de\s+(\d{4})", l)
        if m:
            dia = m.group(1).zfill(2)
            meses = {"enero":"01","febrero":"02","marzo":"03","abril":"04","mayo":"05",
                     "junio":"06","julio":"07","agosto":"08","septiembre":"09",
                     "octubre":"10","noviembre":"11","diciembre":"12"}
            mes = meses.get(m.group(2).lower(), "??")
            plazo = f"{dia}/{mes}"
            break

    if etapa_activa or plazo != "X":
        return "Con cronograma", etapa_activa, plazo
    return "Manifestacion Abierta", "", "X"


# ═══════════════════════════════════════════════════════════════════════════════
# 3. DETECCION DE CAMBIOS
# ═══════════════════════════════════════════════════════════════════════════════

CAMPOS_COMPARAR = ["nombre","direccion","tipo","valor","estado_crono","etapa_actual","plazo"]

def cargar_datos_anteriores():
    if os.path.exists(ARCHIVO_DATOS_PREV):
        with open(ARCHIVO_DATOS_PREV, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def cargar_cambios():
    if os.path.exists(ARCHIVO_CAMBIOS):
        with open(ARCHIVO_CAMBIOS, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def guardar_datos(datos):
    with open(ARCHIVO_DATOS_PREV, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)


def guardar_cambios(cambios):
    with open(ARCHIVO_CAMBIOS, "w", encoding="utf-8") as f:
        json.dump(cambios, f, ensure_ascii=False, indent=2)


def detectar_cambios(inmuebles_nuevos, pestaña):
    """
    Compara datos nuevos con anteriores. Retorna:
    - cambios_activos: dict { "pestaña:id:campo": True } para celdas en rojo
    - resumen_cambios: lista de strings describiendo cada cambio
    """
    anteriores = cargar_datos_anteriores()
    registro   = cargar_cambios()
    ahora      = datetime.now().isoformat()
    limite     = (datetime.now() - timedelta(days=DIAS_ROJO)).isoformat()

    resumen = []

    ids_actuales = set()

    for item in inmuebles_nuevos:
        pid = str(item.get("_id", ""))
        clave_base = f"{pestaña}:{pid}"
        ids_actuales.add(clave_base)
        prev = anteriores.get(clave_base, {})

        # Inmueble NUEVO
        if not prev:
            resumen.append(
                f"NUEVO [{pestaña.upper()}] {item.get('nombre','?')} - "
                f"{item.get('tipo','')} - {item.get('valor','')}"
            )
            for campo in CAMPOS_COMPARAR:
                registro[f"{clave_base}:{campo}"] = ahora

        # Cambios en campos
        for campo in CAMPOS_COMPARAR:
            val_nuevo = str(item.get(campo, ""))
            val_viejo = str(prev.get(campo, ""))

            if prev and val_nuevo != val_viejo:
                clave = f"{clave_base}:{campo}"
                registro[clave] = ahora
                resumen.append(
                    f"CAMBIO [{pestaña.upper()}] {item.get('nombre','?')}: "
                    f"{campo} cambio de '{val_viejo}' a '{val_nuevo}'"
                )

        # Guardar dato actual
        datos_item = {c: str(item.get(c,"")) for c in CAMPOS_COMPARAR}
        anteriores[clave_base] = datos_item

    # Detectar inmuebles ELIMINADOS
    for clave_vieja in list(anteriores.keys()):
        if clave_vieja.startswith(f"{pestaña}:") and clave_vieja not in ids_actuales:
            nombre_viejo = anteriores[clave_vieja].get("nombre", "?")
            tipo_viejo = anteriores[clave_vieja].get("tipo", "")
            resumen.append(f"ELIMINADO [{pestaña.upper()}] {nombre_viejo} - {tipo_viejo}")
            del anteriores[clave_vieja]

    # Limpiar cambios viejos (> DIAS_ROJO)
    for k in list(registro.keys()):
        if registro[k] < limite:
            del registro[k]

    # Armar dict de cambios activos para el Excel
    cambios_activos = {}
    for k, ts in registro.items():
        if ts >= limite:
            cambios_activos[k] = True

    guardar_datos(anteriores)
    guardar_cambios(registro)

    return cambios_activos, resumen


# ═══════════════════════════════════════════════════════════════════════════════
# 4. NOTIFICACION POR EMAIL
# ═══════════════════════════════════════════════════════════════════════════════

def enviar_email(resumen_cambios):
    if not EMAIL_ACTIVADO or not resumen_cambios:
        return

    try:
        nuevos = [c for c in resumen_cambios if c.startswith("NUEVO")]
        eliminados = [c for c in resumen_cambios if c.startswith("ELIMINADO")]
        cambios = [c for c in resumen_cambios if c.startswith("CAMBIO")]

        cuerpo = f"ALERTA - Activos por Colombia\n"
        cuerpo += f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        cuerpo += f"{'='*50}\n\n"

        if nuevos:
            cuerpo += f"INMUEBLES NUEVOS ({len(nuevos)}):\n"
            for c in nuevos: cuerpo += f"  + {c}\n"
            cuerpo += "\n"
        if eliminados:
            cuerpo += f"INMUEBLES ELIMINADOS ({len(eliminados)}):\n"
            for c in eliminados: cuerpo += f"  - {c}\n"
            cuerpo += "\n"
        if cambios:
            cuerpo += f"DATOS QUE CAMBIARON ({len(cambios)}):\n"
            for c in cambios[:40]: cuerpo += f"  * {c}\n"
            if len(cambios) > 40: cuerpo += f"  ... y {len(cambios)-40} mas\n"
            cuerpo += "\n"

        cuerpo += f"{'='*50}\n"
        cuerpo += "Tabla actualizada en Google Drive: inmuebles_medellin.xlsx\n"

        partes = []
        if nuevos: partes.append(f"{len(nuevos)} nuevos")
        if eliminados: partes.append(f"{len(eliminados)} eliminados")
        if cambios: partes.append(f"{len(cambios)} cambios")
        asunto = f"Activos Colombia - {', '.join(partes)}"

        msg = MIMEMultipart()
        msg["From"]    = EMAIL_REMITENTE
        msg["To"]      = EMAIL_DESTINATARIO
        msg["Subject"] = asunto
        msg.attach(MIMEText(cuerpo, "plain"))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_REMITENTE, EMAIL_CONTRASENA)
            server.send_message(msg)
        print(f"  Email enviado a {EMAIL_DESTINATARIO}")
    except Exception as e:
        print(f"  Error enviando email: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# 5. EXCEL CON DOS PESTANAS Y CAMBIOS EN ROJO
# ═══════════════════════════════════════════════════════════════════════════════

def formatear_precio(v):
    if not v: return "X"
    try:    return "$ {:,.0f}".format(float(v)).replace(",",".")
    except: return str(v)


def color_fila(estado_crono, estado_api):
    cod = (estado_api or "").upper()
    if estado_crono == "Manifestacion Abierta": return COLOR_MANIF
    if "SUBASTA" in cod or "PROXIMO" in cod:    return COLOR_SUBASTA
    return COLOR_CRONOGRAMA


def escribir_pestaña(wb, titulo_hoja, titulo_texto, inmuebles, cambios_activos, pestaña):
    ws = wb.create_sheet(titulo_hoja)

    COLS = ["NOMBRE","DIRECCION","TIPO","AREA m2","VALOR",
            "ESTADO CRONOGRAMA","ETAPA ACTUAL","PLAZO","CLIENTE","LINK"]
    CAMPOS = ["nombre","direccion","tipo","area_m2","valor",
              "estado_crono","etapa_actual","plazo","_cliente","link"]
    ANCHOS = [30, 38, 18, 10, 20, 22, 35, 10, 14, 55]
    ALIN   = ["l","l","c","c","c","c","l","c","c","l"]

    borde = openpyxl.styles.Border(
        left=openpyxl.styles.Side("thin", color="CCCCCC"),
        right=openpyxl.styles.Side("thin", color="CCCCCC"),
        top=openpyxl.styles.Side("thin", color="CCCCCC"),
        bottom=openpyxl.styles.Side("thin", color="CCCCCC"),
    )

    # Titulo
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    t = ws.cell(row=1, column=1, value=titulo_texto)
    t.font = openpyxl.styles.Font(name="Calibri", bold=True, size=14, color=FONT_HEADER)
    t.fill = openpyxl.styles.PatternFill("solid", fgColor=COLOR_HEADER)
    t.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Encabezados
    for c, nombre in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=c, value=nombre)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="2E5D9E")
        cell.font = openpyxl.styles.Font(name="Calibri", bold=True, color=FONT_HEADER, size=10)
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde
    ws.row_dimensions[2].height = 30

    # Datos
    relleno_rojo = openpyxl.styles.PatternFill("solid", fgColor=COLOR_CAMBIO)
    fuente_rojo  = openpyxl.styles.Font(name="Calibri", size=10, color="FFFFFF", bold=True)

    for fila, item in enumerate(inmuebles, 3):
        pid = str(item.get("_id",""))
        color_base = color_fila(item.get("estado_crono",""), item.get("estado_api",""))
        relleno_base = openpyxl.styles.PatternFill("solid", fgColor=color_base)
        fuente_base  = openpyxl.styles.Font(name="Calibri", size=10)

        for c, (campo, alin) in enumerate(zip(CAMPOS, ALIN), 1):
            val = "Grupo NBC" if campo == "_cliente" else item.get(campo, "")
            cell = ws.cell(row=fila, column=c, value=val)

            # Verificar si este campo cambio (rojo)
            clave_cambio = f"{pestaña}:{pid}:{campo}"
            if clave_cambio in cambios_activos:
                cell.fill = relleno_rojo
                cell.font = fuente_rojo
            else:
                cell.fill = relleno_base
                cell.font = fuente_base

            cell.alignment = openpyxl.styles.Alignment(
                horizontal="left" if alin == "l" else "center",
                vertical="center", wrap_text=(alin == "l")
            )
            cell.border = borde
        ws.row_dimensions[fila].height = 22

    # Anchos y filtros
    for c, w in enumerate(ANCHOS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.auto_filter.ref = f"A2:{openpyxl.utils.get_column_letter(len(COLS))}{len(inmuebles)+2}"
    ws.freeze_panes = "A3"


def guardar_excel(medellin, antioquia, cambios_med, cambios_ant):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # eliminar hoja por defecto

    escribir_pestaña(wb, "Medellin", "LISTADO DE VENTA MASIVA - MEDELLIN",
                     medellin, cambios_med, "med")
    escribir_pestaña(wb, "Antioquia", "LISTADO DE VENTA MASIVA - ANTIOQUIA (sin Medellin)",
                     antioquia, cambios_ant, "ant")

    # Leyenda
    ws = wb.create_sheet("Leyenda")
    ws["A1"] = "Leyenda de colores"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=12)
    leyenda = [
        ("Con cronograma - Proximo Subasta", COLOR_SUBASTA),
        ("Con cronograma - En proceso",      COLOR_CRONOGRAMA),
        ("Manifestacion Abierta",            COLOR_MANIF),
        ("DATO QUE CAMBIO (rojo por 2 dias)", COLOR_CAMBIO),
    ]
    for i, (desc, col) in enumerate(leyenda, 2):
        c = ws.cell(row=i, column=1, value=desc)
        c.fill = openpyxl.styles.PatternFill("solid", fgColor=col)
        if col == COLOR_CAMBIO:
            c.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 50

    # Info
    ws2 = wb.create_sheet("Info")
    ws2["A1"] = "Ultima actualizacion";  ws2["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2["A2"] = "Medellin";             ws2["B2"] = f"{len(medellin)} inmuebles"
    ws2["A3"] = "Antioquia (sin Med.)"; ws2["B3"] = f"{len(antioquia)} inmuebles"
    ws2["A4"] = "Fuente";               ws2["B4"] = "activosporcolombia.com"

    wb.save(ARCHIVO_EXCEL)
    print(f"Guardado: {ARCHIVO_EXCEL}")


# ═══════════════════════════════════════════════════════════════════════════════
# 6. ARMAR DATOS FINALES
# ═══════════════════════════════════════════════════════════════════════════════

def procesar_propiedades(props_api, detalles_scrape):
    """Combina datos de la API con los detalles scrapeados."""
    resultado = []
    for prop in props_api:
        pid = str(prop["id"])
        det = detalles_scrape.get(pid, {})

        estado_api = prop.get("state") or {}
        estado_cod = estado_api.get("code","") if isinstance(estado_api, dict) else ""
        tipo_id = prop.get("property_type_id")

        resultado.append({
            "_id":          pid,
            "nombre":       det.get("nombre") or det.get("barrio") or prop.get("reference",""),
            "direccion":    det.get("direccion",""),
            "tipo":         TIPOS.get(tipo_id, f"Tipo {tipo_id}" if tipo_id else ""),
            "area_m2":      prop.get("built_area") or prop.get("lot_area") or "",
            "valor":        formatear_precio(prop.get("base_sale_price") or prop.get("commercial_appraisal")),
            "estado_crono": det.get("estado_crono",""),
            "etapa_actual": det.get("etapa_actual",""),
            "plazo":        det.get("plazo","X"),
            "estado_api":   estado_cod,
            "link":         construir_url(prop),
        })
    return resultado


# ═══════════════════════════════════════════════════════════════════════════════
# 7. MAIN
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 55)
    print("  Scraping activosporcolombia.com")
    print("  Medellin + Antioquia")
    print("=" * 55)

    # ── Descargar listas ──
    print("\n[1/5] Descargando propiedades de Medellin...")
    props_med = obtener_medellin()
    print(f"  {len(props_med)} propiedades")

    print("\n[2/5] Descargando propiedades de Antioquia (sin Medellin)...")
    props_ant = obtener_antioquia_sin_medellin()
    print(f"  {len(props_ant)} propiedades")

    # ── Scrape detalles ──
    print("\n[3/5] Visitando paginas de Medellin...")
    det_med = scrape_detalles(props_med, "MED ")

    print("\n[4/5] Visitando paginas de Antioquia...")
    det_ant = scrape_detalles(props_ant, "ANT ")

    # ── Procesar ──
    inmuebles_med = procesar_propiedades(props_med, det_med)
    inmuebles_ant = procesar_propiedades(props_ant, det_ant)

    # ── Detectar cambios ──
    print("\n[5/5] Detectando cambios y generando Excel...")
    cambios_med, resumen_med = detectar_cambios(inmuebles_med, "med")
    cambios_ant, resumen_ant = detectar_cambios(inmuebles_ant, "ant")
    todos_cambios = resumen_med + resumen_ant

    if todos_cambios:
        print(f"\n  *** {len(todos_cambios)} CAMBIOS DETECTADOS ***")
        for c in todos_cambios[:15]:
            print(f"    - {c}")
        if len(todos_cambios) > 15:
            print(f"    ... y {len(todos_cambios)-15} mas")
    else:
        print("  Sin cambios respecto a la ultima actualizacion")

    # ── Excel ──
    guardar_excel(inmuebles_med, inmuebles_ant, cambios_med, cambios_ant)

    # ── Copiar a Drive ──
    try:
        shutil.copy2(ARCHIVO_EXCEL, ARCHIVO_DRIVE)
        print(f"Copiado a Drive: {ARCHIVO_DRIVE}")
    except Exception as e:
        print(f"Drive: {e}")

    # ── Notificacion ──
    if todos_cambios:
        enviar_email(todos_cambios)

    print(f"\nListo! {len(inmuebles_med)} Medellin + {len(inmuebles_ant)} Antioquia")
